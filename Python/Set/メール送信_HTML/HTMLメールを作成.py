#!/usr/bin/env python
# coding: utf-8

# In[1]:


#outlookは起動しておいてください
from win32com.client import Dispatch
import datetime 
import win32com.client
import openpyxl


#htmlファイルを読み込み
filepath = 'HTML本文.html'
with open(filepath , encoding='utf-8') as f:
    html = f.read()

#メール配信リスト列名取り込み
wb = openpyxl.load_workbook('メール配信リスト.xlsx')
sheet = wb['sheet1']
cell = sheet.cell(row=2, column=1)

#メール件名
f = open('メール件名.txt')
mailsubject = f.read() 
f.close()

#outlook設定
outlook = Dispatch("Outlook.Application").GetNamespace("MAPI")
inbox = outlook.GetDefaultFolder("6")
const=win32com.client.constants
olMailItem = 0x0
obj = win32com.client.Dispatch("Outlook.Application")

#エクセルの列名カウント
i = 1

while True:
    i = i + 1
    cell = sheet.cell(row=i, column=1)
    mailadress = cell.value
    
    if mailadress != None:
        #メール送信
        newMail = obj.CreateItem(olMailItem)
        newMail.Subject = mailsubject
        newMail.HtmlBody = html
        newMail.To = mailadress
        newMail.display()
        newMail.Send()
        print (mailadress)
    else:
        break


# In[ ]:




