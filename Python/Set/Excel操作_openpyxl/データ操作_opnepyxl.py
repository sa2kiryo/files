#!/usr/bin/env python
# coding: utf-8

# In[32]:


#使用するライブラリをインポート
import openpyxl


#エクセル取り込み
wb = openpyxl.load_workbook('order.xlsx')
sheet = wb['Sheet1']

#エクセルの列名カウント
i = 1

#エクセルの縦1、横3行目に「進捗」と書き込む。
sheet.cell(row=1, column=3).value = "電話番号_完" 
sheet.cell(row=1, column=4).value = "進捗" 

#ループ開始
while True:
    i = i + 1
    
    #注文番号の取得
    cell_c = sheet.cell(row=i, column=1)
    entity_id = cell_c.value

    #注文番号が存在していれば、次の操作を行う
    if entity_id != None:
        
        #▽▽▽▽▽▽▽▽電話番号操作▽▽▽▽▽▽▽▽
        
        #電話番号を取得
        cell_t = sheet.cell(row=i, column=2)
        
        #電話番号を文字列として変換しておく（これやらないと、数値と認識されて、文字操作ができなくなる）
        tel_id = str(cell_t.value)        
        
        #文字置き換え（ハイフンが含まれたら、消す）
        tel_id = tel_id.replace('-', '')
        
        #加工した電話番号を、もとのエクセルに書き込みなおす
        sheet.cell(row=i, column=3).value = tel_id
        
        #△△△△△△△△電話番号操作終了△△△△△△△△
        
        
        #対応した列に「対応済み」と記載
        sheet.cell(row=i, column=4).value = str(i) + "番目OK"
        
    else:
        #注文番号がなければ、ループを抜ける。
        break
        
wb.save('order_comp.xlsx')

print("完了")


# In[ ]:




