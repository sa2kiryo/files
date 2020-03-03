#!/usr/bin/env python
# coding: utf-8

# In[22]:


import openpyxl
import os


#画像名を変更したいリスト列名取り込み
#1列目に変更前の画像名、二列目に変更後の画像名を入力
wb = openpyxl.load_workbook('画像名変更希望リスト.xlsx')
sheet = wb['Sheet1']



#エクセルの列カウント
i = 1

while True:
    i = i + 1
    
    #変更前の画像名を取得
    cell = sheet.cell(row=i, column=1)
    image1 = cell.value
    path1 = r"C:\temp\img\" + str(image1) + ".jpg"
    
    #変更後の画像名を取得
    cell2 = sheet.cell(row=i, column=2)
    image2 = cell2.value
    path2 = r"C:\temp\img_ok\" + str(image2) + ".jpg"    
    
    
    if image1 != None:
        #確認
        print (path1)
        print (image2)
        
        #画像名を変更
        os.rename(path1, path2) 

    else:
        break


# In[ ]:




